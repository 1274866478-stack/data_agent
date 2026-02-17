# -*- coding: utf-8 -*-
"""
测试 Excel 数据库生成脚本

生成符合表关系的测试数据，输出为 Excel 文件。
每个 sheet 对应一个表，按依赖顺序排列。

运行方式:
    python scripts/generate_test_database.py
"""

import random
import sys
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

# Windows 下设置 UTF-8 输出
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# 测试数据源
# ============================================================================

# 中国省市区数据
PROVINCES = [
    "北京", "上海", "广东", "浙江", "江苏", "安徽", "山东", "河南", "四川", "湖北"
]

CITIES = {
    "北京": ["东城", "西城", "朝阳", "海淀", "丰台"],
    "上海": ["黄浦", "徐汇", "静安", "浦东", "闵行"],
    "广东": ["广州", "深圳", "佛山", "东莞", "珠海"],
    "浙江": ["杭州", "宁波", "温州", "嘉兴", "绍兴"],
    "江苏": ["南京", "苏州", "无锡", "常州", "徐州"],
    "安徽": ["合肥", "芜湖", "蚌埠", "淮南", "马鞍山"],
    "山东": ["济南", "青岛", "淄博", "烟台", "潍坊"],
    "河南": ["郑州", "开封", "洛阳", "平顶山", "安阳"],
    "四川": ["成都", "绵阳", "自贡", "攀枝花", "泸州"],
    "湖北": ["武汉", "黄石", "十堰", "宜昌", "襄阳"]
}

DISTRICTS = [
    "朝阳区", "海淀区", "西城区", "东城区", "丰台区",
    "黄浦区", "徐汇区", "静安区", "浦东新区", "闵行区",
    "天河区", "越秀区", "海珠区", "白云区", "番禺区",
    "福田区", "南山区", "罗湖区", "宝安区", "龙岗区",
    "西湖区", "拱墅区", "滨江区", "余杭区", "萧山区"
]

# 产品分类
CATEGORIES_DATA = [
    (1, "电子产品"),
    (2, "家居用品"),
    (3, "服装鞋帽"),
    (4, "食品饮料"),
    (5, "图书文具")
]

# 产品数据
PRODUCTS_DATA = [
    # 电子产品 (category_id = 1)
    (1, "iPhone 15 Pro", Decimal("7999.00"), 1),
    (2, "MacBook Air M2", Decimal("8999.00"), 1),
    (3, "AirPods Pro 2", Decimal("1899.00"), 1),
    # 家居用品 (category_id = 2)
    (4, "北欧风布艺沙发", Decimal("2999.00"), 2),
    (5, "实木餐桌", Decimal("1599.00"), 2),
    (6, "智能扫地机器人", Decimal("1299.00"), 2),
    # 服装鞋帽 (category_id = 3)
    (7, "男士商务衬衫", Decimal("299.00"), 3),
    (8, "女士连衣裙", Decimal("399.00"), 3),
    (9, "运动跑步鞋", Decimal("599.00"), 3),
    # 食品饮料 (category_id = 4)
    (10, "有机茶叶礼盒", Decimal("199.00"), 4),
    (11, "进口红酒", Decimal("399.00"), 4),
    (12, "坚果礼盒", Decimal("129.00"), 4),
    # 图书文具 (category_id = 5)
    (13, "Python编程从入门到精通", Decimal("89.00"), 5),
    (14, "高级钢笔礼盒", Decimal("299.00"), 5),
    (15, "办公笔记本套装", Decimal("59.00"), 5),
]

# 用户数据
USERS_DATA = [
    (1, "张伟", "zhangwei@email.com", "13800138001", "2023-01-15"),
    (2, "李娜", "lina@email.com", "13800138002", "2023-02-20"),
    (3, "王芳", "wangfang@email.com", "13800138003", "2023-03-10"),
    (4, "刘洋", "liuyang@email.com", "13800138004", "2023-03-25"),
    (5, "陈静", "chenjing@email.com", "13800138005", "2023-04-12"),
    (6, "杨明", "yangming@email.com", "13800138006", "2023-05-08"),
    (7, "赵强", "zhaoqiang@email.com", "13800138007", "2023-06-18"),
    (8, "黄婷", "huangting@email.com", "13800138008", "2023-07-22"),
    (9, "周杰", "zhoujie@email.com", "13800138009", "2023-08-30"),
    (10, "吴磊", "wulei@email.com", "13800138010", "2023-09-15"),
]

# 详细地址
DETAIL_ADDRESSES = [
    "建设路1号院3栋201室",
    "人民路88号花园小区5栋302",
    "中山大道123号阳光公寓2单元801",
    "解放路56号和平里小区6栋101",
    "新华路99号金域华庭8栋1501",
    "长安街168号城市花园3栋502",
    "复兴路200号保利家园1栋1201",
    "文化路66号书香门第4栋701",
    "科技路188号创新大厦A座2001",
    "花园路88号玫瑰园9栋301",
    "体育路50号奥林匹克花园2栋601",
    "教育路120号学府苑5栋1001",
    "商业路180号万达广场公寓3栋401",
    "工业路90号产业园职工宿舍1栋201",
    "友谊路66号友谊小区7栋901",
]


# ============================================================================
# 数据生成器类
# ============================================================================

class TestDatabaseGenerator:
    """测试数据库生成器"""

    def __init__(self, output_path: str | None = None):
        """
        初始化生成器

        Args:
            output_path: 输出文件路径，默认为 scripts/test_database.xlsx
        """
        if output_path is None:
            script_dir = Path(__file__).parent
            output_path = str(script_dir / "test_database.xlsx")

        self.output_path = Path(output_path)
        self.wb = openpyxl.Workbook()
        # 删除默认sheet
        if self.wb.active is not None:
            self.wb.remove(self.wb.active)

        # 样式定义
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # 存储生成的ID，确保外键引用正确
        self.category_ids: list[int] = []
        self.product_ids: list[int] = []
        self.user_ids: list[int] = []
        self.address_ids: list[int] = []
        self.order_ids: list[int] = []
        self.order_item_count: int = 0  # 订单明细数量

    def add_sheet(self, sheet_name: str, columns: list, data: list):
        """
        添加工作表

        Args:
            sheet_name: 工作表名称
            columns: 列名列表
            data: 数据列表（每行是一个tuple）
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        print(f"✓ 创建工作表 '{sheet_name}': {len(data)} 条记录")

    def generate_categories(self):
        """生成分类表"""
        columns = ["id", "name"]
        data = list(CATEGORIES_DATA)
        self.category_ids = [row[0] for row in data]
        self.add_sheet("categories", columns, data)

    def generate_products(self):
        """生成产品表"""
        columns = ["id", "name", "price", "category_id"]
        data = list(PRODUCTS_DATA)
        self.product_ids = [row[0] for row in data]
        self.add_sheet("products", columns, data)

    def generate_users(self):
        """生成用户表"""
        columns = ["id", "username", "email", "phone", "registration_date"]
        data = list(USERS_DATA)
        self.user_ids = [row[0] for row in data]
        self.add_sheet("users", columns, data)

    def generate_addresses(self):
        """生成地址表"""
        columns = ["id", "user_id", "province", "city", "district", "detail_address", "is_default"]

        data = []
        address_id = 1

        # 为每个用户生成1-2个地址
        for user_id in self.user_ids:
            # 每个用户至少1个地址，50%概率有2个地址
            num_addresses = 2 if random.random() > 0.5 else 1

            for addr_idx in range(num_addresses):
                province = random.choice(PROVINCES)
                city = random.choice(CITIES[province])
                district = random.choice(DISTRICTS)
                detail_address = DETAIL_ADDRESSES[(address_id - 1) % len(DETAIL_ADDRESSES)]
                is_default = addr_idx == 0  # 第一个地址为默认地址

                data.append((address_id, user_id, province, city, district, detail_address, is_default))
                self.address_ids.append(address_id)
                address_id += 1

        self.add_sheet("addresses", columns, data)

    def generate_orders(self):
        """生成订单表"""
        columns = ["id", "user_id", "order_date", "total_amount", "created_at"]

        data = []
        order_id = 1
        base_date = datetime(2024, 1, 1)

        # 为每个用户生成1-3个订单
        for user_id in self.user_ids:
            num_orders = random.randint(1, 3)

            for _ in range(num_orders):
                # 随机生成订单日期（2024年1月到2024年12月）
                days_offset = random.randint(0, 364)
                order_date = base_date + timedelta(days=days_offset)

                # 添加时间到 created_at
                hour = random.randint(9, 18)
                minute = random.randint(0, 59)
                created_at = order_date.replace(hour=hour, minute=minute)

                # 随机订单金额
                total_amount = Decimal(str(round(random.uniform(100, 5000), 2)))

                data.append((order_id, user_id, order_date.strftime("%Y-%m-%d"),
                            total_amount, created_at.strftime("%Y-%m-%d %H:%M:%S")))
                self.order_ids.append(order_id)
                order_id += 1

        # 按订单日期排序
        data.sort(key=lambda x: x[2])

        # 更新ID顺序
        for idx, (original_id, *rest) in enumerate(data, 1):
            data[idx - 1] = (idx, *rest)

        self.order_ids = [row[0] for row in data]
        self.add_sheet("orders", columns, data)

    def generate_order_items(self):
        """生成订单明细表"""
        columns = ["id", "order_id", "product_id", "quantity", "unit_price"]

        data = []
        item_id = 1

        # 为每个订单生成1-3个明细
        for order_id in self.order_ids:
            num_items = random.randint(1, 3)
            selected_products = random.sample(self.product_ids, min(num_items, len(self.product_ids)))

            for product_id in selected_products:
                # 获取产品单价
                product = next(p for p in PRODUCTS_DATA if p[0] == product_id)
                unit_price = product[2]
                quantity = random.randint(1, 5)

                data.append((item_id, order_id, product_id, quantity, unit_price))
                item_id += 1

        # 保存订单明细数量供统计使用
        self.order_item_count = item_id - 1
        self.add_sheet("order_items", columns, data)

    def generate(self):
        """生成所有数据表"""
        print("=" * 60)
        print("开始生成测试数据库 Excel 文件")
        print("=" * 60)

        # 按依赖顺序生成
        print("\n[1/6] 生成分类表...")
        self.generate_categories()

        print("\n[2/6] 生成产品表...")
        self.generate_products()

        print("\n[3/6] 生成用户表...")
        self.generate_users()

        print("\n[4/6] 生成地址表...")
        self.generate_addresses()

        print("\n[5/6] 生成订单表...")
        self.generate_orders()

        print("\n[6/6] 生成订单明细表...")
        self.generate_order_items()

        # 保存文件
        print("\n" + "=" * 60)
        self.wb.save(self.output_path)
        print(f"✓ 测试数据库已生成: {self.output_path}")
        print("=" * 60)

        # 打印统计信息
        print("\n📊 数据统计:")
        print(f"  - 分类 (categories): {len(self.category_ids)} 条")
        print(f"  - 产品 (products): {len(self.product_ids)} 条")
        print(f"  - 用户 (users): {len(self.user_ids)} 条")
        print(f"  - 地址 (addresses): {len(self.address_ids)} 条")
        print(f"  - 订单 (orders): {len(self.order_ids)} 条")
        print(f"  - 订单明细 (order_items): {self.order_item_count} 条")

        # 验证外键完整性
        print("\n🔍 验证外键完整性:")
        self._validate_foreign_keys()

    def _validate_foreign_keys(self):
        """验证外键引用完整性"""
        all_valid = True

        # 验证 products.category_id
        for product in PRODUCTS_DATA:
            if product[3] not in self.category_ids:
                print(f"  ✗ products.id={product[0]} 的 category_id={product[3]} 不存在")
                all_valid = False

        # 读取 addresses 验证 user_id
        ws_addresses = self.wb["addresses"]
        for row in list(ws_addresses.iter_rows(min_row=2, values_only=True)):
            if row[1] not in self.user_ids:
                print(f"  ✗ addresses.id={row[0]} 的 user_id={row[1]} 不存在")
                all_valid = False

        # 读取 orders 验证 user_id
        ws_orders = self.wb["orders"]
        for row in list(ws_orders.iter_rows(min_row=2, values_only=True)):
            if row[1] not in self.user_ids:
                print(f"  ✗ orders.id={row[0]} 的 user_id={row[1]} 不存在")
                all_valid = False

        # 读取 order_items 验证 order_id 和 product_id
        ws_order_items = self.wb["order_items"]
        for row in list(ws_order_items.iter_rows(min_row=2, values_only=True)):
            if row[1] not in self.order_ids:
                print(f"  ✗ order_items.id={row[0]} 的 order_id={row[1]} 不存在")
                all_valid = False
            if row[2] not in self.product_ids:
                print(f"  ✗ order_items.id={row[0]} 的 product_id={row[2]} 不存在")
                all_valid = False

        if all_valid:
            print("  ✓ 所有外键引用完整，验证通过！")


# ============================================================================
# 主程序
# ============================================================================

if __name__ == "__main__":
    # 设置随机种子以保证可重复性
    random.seed(42)

    # 创建生成器并生成数据
    generator = TestDatabaseGenerator()
    generator.generate()
